"""
Comprehensive Report Generator for OpenEvolve Decomposition Engine

This module provides advanced reporting capabilities with multiple export formats,
custom templates, automated scheduling, and executive summaries.

Features:
- PDF reports with embedded charts and graphs
- HTML interactive reports with navigation
- Markdown documentation export
- Excel spreadsheet exports with multiple sheets
- JSON data exports for API integration
- Custom templates with Jinja2
- Report aggregation across multiple decompositions
- Executive summaries with AI-powered insights
- Branding and customization support
- Scheduled report generation

Export Formats:
- PDF: Professional documents with charts, tables, and formatting
- HTML: Interactive web-based reports with filtering and sorting
- Markdown: Documentation-friendly format for wikis and READMEs
- Excel: Multi-sheet spreadsheets with raw data and summaries
- JSON: Structured data for API consumption and further processing
"""

import logging
import json
from typing import Dict, List, Any, Optional, Tuple, Union
from datetime import datetime, timedelta
from pathlib import Path
from io import BytesIO
import base64
from collections import defaultdict
import copy

logger = logging.getLogger(__name__)

# Import visualization library
try:
    from progress_visualizer import ProgressVisualizer
    VISUALIZER_AVAILABLE = True
except ImportError:
    logger.warning("ProgressVisualizer not found. Chart generation will be limited.")
    VISUALIZER_AVAILABLE = False

# Import quality tracker
try:
    from quality_tracker import QualityTracker
    QUALITY_TRACKER_AVAILABLE = True
except ImportError:
    logger.warning("QualityTracker not found. Quality trend analysis will be limited.")
    QUALITY_TRACKER_AVAILABLE = False

# Import knowledge base
try:
    from knowledge_base import KnowledgeBase
    KNOWLEDGE_BASE_AVAILABLE = True
except ImportError:
    logger.warning("KnowledgeBase not found. Historical data will be limited.")
    KNOWLEDGE_BASE_AVAILABLE = False

# Import template engine
try:
    from jinja2 import Environment, FileSystemLoader, Template, BaseLoader
    from jinja2 import TemplateNotFound
    JINJA2_AVAILABLE = True
except ImportError:
    logger.warning("Jinja2 not found. Custom templates will not be available.")
    JINJA2_AVAILABLE = False

# Import PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.platypus import Image, KeepTogether
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    logger.warning("ReportLab not found. PDF generation will be disabled.")
    REPORTLAB_AVAILABLE = False

# Import Excel generation
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    logger.warning("openpyxl not found. Excel generation will be disabled.")
    OPENPYXL_AVAILABLE = False

# Import plotting libraries
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    logger.warning("matplotlib not found. Chart generation will be limited.")
    MATPLOTLIB_AVAILABLE = False


class ReportGenerator:
    """
    Comprehensive report generator with multiple export formats.

    Supports:
    - PDF generation with ReportLab
    - HTML interactive reports
    - Markdown documentation
    - Excel spreadsheets
    - JSON data exports
    - Custom templates with Jinja2
    - Report aggregation
    - Executive summaries
    """

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize report generator.

        Args:
            config: Optional configuration dictionary with settings:
                - output_dir: Base directory for report output (default: ./reports)
                - template_dir: Directory for custom templates (default: ./templates)
                - branding: Dict with branding options (logo, colors, fonts)
                - default_format: Default export format (default: html)
                - include_charts: Whether to include charts (default: True)
                - include_raw_data: Whether to include raw data (default: False)
        """
        self.config = config or {}
        self.output_dir = Path(self.config.get('output_dir', './reports'))
        self.template_dir = Path(self.config.get('template_dir', './templates'))
        self.branding = self.config.get('branding', {})
        self.default_format = self.config.get('default_format', 'html')
        self.include_charts = self.config.get('include_charts', True)
        self.include_raw_data = self.config.get('include_raw_data', False)

        # Create output directory
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Initialize components
        self.visualizer = ProgressVisualizer() if VISUALIZER_AVAILABLE else None
        self.quality_tracker = QualityTracker() if QUALITY_TRACKER_AVAILABLE else None
        self.knowledge_base = KnowledgeBase() if KNOWLEDGE_BASE_AVAILABLE else None

        # Initialize Jinja2 environment if available
        self.jinja_env = None
        if JINJA2_AVAILABLE and self.template_dir.exists():
            try:
                self.jinja_env = Environment(
                    loader=FileSystemLoader(str(self.template_dir)),
                    autoescape=True
                )
                logger.info(f"Jinja2 environment initialized with template dir: {self.template_dir}")
            except (OSError, IOError, ImportError, ValueError) as e:
                logger.warning(f"Failed to initialize Jinja2 environment: {e}")

        # Track available formats
        self.available_formats = ['html', 'json', 'md']
        if REPORTLAB_AVAILABLE:
            self.available_formats.append('pdf')
        if OPENPYXL_AVAILABLE:
            self.available_formats.append('xlsx')

        logger.info(f"ReportGenerator initialized. Available formats: {self.available_formats}")

    def generate_report(
        self,
        decomposition_plan: Any,
        output_format: str = None,
        template_name: Optional[str] = None,
        output_path: Optional[str] = None,
        include_sections: Optional[List[str]] = None,
        custom_data: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        Generate a comprehensive report in the specified format.

        Args:
            decomposition_plan: DecompositionPlan object or dict
            output_format: Format for output (pdf, html, md, xlsx, json)
            template_name: Optional custom template name
            output_path: Optional path for output file
            include_sections: List of sections to include (default: all)
            custom_data: Optional custom data to include in report

        Returns:
            str: Path to generated report file

        Raises:
            ValueError: If format is not supported
        """
        output_format = output_format or self.default_format

        if output_format not in self.available_formats:
            raise ValueError(f"Unsupported format: {output_format}. Available: {self.available_formats}")

        logger.info(f"Generating {output_format} report for decomposition plan")

        # Convert plan to dict if needed
        plan_data = self._extract_plan_data(decomposition_plan)

        # Merge custom data
        if custom_data:
            plan_data.update(custom_data)

        # Determine sections to include
        sections = include_sections or self._get_default_sections()

        # Add metadata
        plan_data['metadata'] = self._generate_metadata(plan_data)
        plan_data['report_config'] = {
            'format': output_format,
            'generated_at': datetime.now().isoformat(),
            'sections': sections,
            'branding': self.branding
        }

        # Generate report based on format
        if output_format == 'pdf':
            return self._generate_pdf_report(plan_data, template_name, output_path, sections)
        elif output_format == 'html':
            return self._generate_html_report(plan_data, template_name, output_path, sections)
        elif output_format == 'md':
            return self._generate_markdown_report(plan_data, template_name, output_path, sections)
        elif output_format == 'xlsx':
            return self._generate_excel_report(plan_data, output_path, sections)
        elif output_format == 'json':
            return self._generate_json_report(plan_data, output_path)
        else:
            raise ValueError(f"Unsupported format: {output_format}")

    def generate_aggregated_report(
        self,
        decomposition_plans: List[Any],
        output_format: str = 'html',
        group_by: str = 'date',
        output_path: Optional[str] = None
    ) -> str:
        """
        Generate an aggregated report across multiple decomposition plans.

        Args:
            decomposition_plans: List of DecompositionPlan objects
            output_format: Format for output
            group_by: How to group data (date, strategy, problem_type)
            output_path: Optional path for output file

        Returns:
            str: Path to generated report
        """
        logger.info(f"Generating aggregated report for {len(decomposition_plans)} plans")

        # Extract data from all plans
        plans_data = [self._extract_plan_data(plan) for plan in decomposition_plans]

        # Group and aggregate data
        aggregated_data = self._aggregate_data(plans_data, group_by)

        # Add aggregation metadata
        aggregated_data['metadata'] = self._generate_metadata(aggregated_data)
        aggregated_data['aggregation'] = {
            'total_plans': len(decomposition_plans),
            'group_by': group_by,
            'generated_at': datetime.now().isoformat()
        }

        # Generate report
        if output_format == 'pdf':
            return self._generate_pdf_report(aggregated_data, None, output_path, ['aggregated'])
        elif output_format == 'html':
            return self._generate_html_report(aggregated_data, None, output_path, ['aggregated'])
        elif output_format == 'md':
            return self._generate_markdown_report(aggregated_data, None, output_path, ['aggregated'])
        elif output_format == 'xlsx':
            return self._generate_excel_report(aggregated_data, output_path, ['aggregated'])
        elif output_format == 'json':
            return self._generate_json_report(aggregated_data, output_path)
        else:
            raise ValueError(f"Unsupported format: {output_format}")

    def generate_executive_summary(
        self,
        decomposition_plan: Any,
        output_format: str = 'pdf',
        output_path: Optional[str] = None
    ) -> str:
        """
        Generate an executive summary with key insights and metrics.

        Args:
            decomposition_plan: DecompositionPlan object
            output_format: Format for output
            output_path: Optional path for output file

        Returns:
            str: Path to generated report
        """
        logger.info("Generating executive summary")

        plan_data = self._extract_plan_data(decomposition_plan)

        # Generate executive summary data
        summary_data = {
            'title': 'Executive Summary',
            'problem': plan_data.get('problem', {}),
            'key_metrics': self._calculate_key_metrics(plan_data),
            'insights': self._generate_insights(plan_data),
            'recommendations': self._generate_recommendations(plan_data),
            'quality_summary': self._summarize_quality(plan_data),
            'timeline': self._summarize_timeline(plan_data),
            'metadata': self._generate_metadata(plan_data)
        }

        # Generate report
        if output_format == 'pdf':
            return self._generate_pdf_report(summary_data, 'executive_summary', output_path, ['executive'])
        elif output_format == 'html':
            return self._generate_html_report(summary_data, 'executive_summary', output_path, ['executive'])
        elif output_format == 'md':
            return self._generate_markdown_report(summary_data, 'executive_summary', output_path, ['executive'])
        else:
            raise ValueError(f"Executive summary not available in format: {output_format}")

    def _extract_plan_data(self, plan: Any) -> Dict[str, Any]:
        """Extract data from decomposition plan to dict."""
        if hasattr(plan, '__dict__'):
            return copy.deepcopy(plan.__dict__)
        elif isinstance(plan, dict):
            return copy.deepcopy(plan)
        else:
            return {'raw': plan}

    def _generate_metadata(self, plan_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate report metadata."""
        return {
            'generated_at': datetime.now().isoformat(),
            'generator_version': '1.0.0',
            'plan_id': plan_data.get('plan_id', 'unknown'),
            'problem_title': plan_data.get('problem', {}).get('title', 'Unknown'),
            'strategy': plan_data.get('strategy', 'unknown'),
            'total_sub_problems': len(plan_data.get('sub_problems', [])),
            'completion_percentage': self._calculate_completion(plan_data)
        }

    def _calculate_completion(self, plan_data: Dict[str, Any]) -> float:
        """Calculate completion percentage."""
        sub_problems = plan_data.get('sub_problems', [])
        if not sub_problems:
            return 0.0

        completed = sum(1 for sp in sub_problems if sp.get('status') == 'completed')
        return (completed / len(sub_problems)) * 100

    def _calculate_key_metrics(self, plan_data: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate key metrics for executive summary."""
        sub_problems = plan_data.get('sub_problems', [])

        return {
            'total_sub_problems': len(sub_problems),
            'completed': sum(1 for sp in sub_problems if sp.get('status') == 'completed'),
            'in_progress': sum(1 for sp in sub_problems if sp.get('status') == 'in_progress'),
            'pending': sum(1 for sp in sub_problems if sp.get('status') == 'pending'),
            'blocked': sum(1 for sp in sub_problems if sp.get('status') == 'blocked'),
            'completion_percentage': self._calculate_completion(plan_data),
            'average_quality_score': self._calculate_average_quality(plan_data),
            'total_dependencies': sum(len(sp.get('dependencies', [])) for sp in sub_problems),
            'critical_path_length': self._calculate_critical_path_length(plan_data),
            'estimated_completion': self._estimate_completion(plan_data)
        }

    def _calculate_average_quality(self, plan_data: Dict[str, Any]) -> float:
        """Calculate average quality score."""
        sub_problems = plan_data.get('sub_problems', [])
        quality_scores = [sp.get('quality_score', 0) for sp in sub_problems if sp.get('quality_score')]

        if not quality_scores:
            return 0.0

        return sum(quality_scores) / len(quality_scores)

    def _calculate_critical_path_length(self, plan_data: Dict[str, Any]) -> int:
        """Calculate critical path length."""
        # Simplified critical path calculation
        sub_problems = plan_data.get('sub_problems', [])
        return max(len(sp.get('dependencies', [])) for sp in sub_problems) if sub_problems else 0

    def _estimate_completion(self, plan_data: Dict[str, Any]) -> str:
        """Estimate completion date."""
        # Simplified estimation - would need more sophisticated logic
        return 'Not available'

    def _generate_insights(self, plan_data: Dict[str, Any]) -> List[str]:
        """Generate AI-powered insights."""
        insights = []

        # Analyze completion rate
        completion = self._calculate_completion(plan_data)
        if completion > 80:
            insights.append("Project is nearing completion with strong progress")
        elif completion > 50:
            insights.append("Project is past halfway point with steady progress")
        elif completion > 20:
            insights.append("Project is in early stages with initial momentum")
        else:
            insights.append("Project is in initial setup phase")

        # Analyze bottlenecks
        sub_problems = plan_data.get('sub_problems', [])
        blocked = [sp for sp in sub_problems if sp.get('status') == 'blocked']
        if blocked:
            insights.append(f"Attention needed: {len(blocked)} sub-problems are currently blocked")

        # Analyze dependencies
        high_deps = [sp for sp in sub_problems if len(sp.get('dependencies', [])) > 3]
        if high_deps:
            insights.append(f"Complex dependencies: {len(high_deps)} sub-problems have 3+ dependencies")

        return insights

    def _generate_recommendations(self, plan_data: Dict[str, Any]) -> List[str]:
        """Generate actionable recommendations."""
        recommendations = []

        # Check for blocked items
        sub_problems = plan_data.get('sub_problems', [])
        blocked = [sp for sp in sub_problems if sp.get('status') == 'blocked']
        if blocked:
            recommendations.append("Address blocked sub-problems to prevent delays")

        # Check for overdue items
        # (would need date comparison logic)

        # Check quality scores
        low_quality = [sp for sp in sub_problems if sp.get('quality_score', 100) < 70]
        if low_quality:
            recommendations.append("Review low-quality sub-problems for improvement opportunities")

        return recommendations

    def _summarize_quality(self, plan_data: Dict[str, Any]) -> Dict[str, Any]:
        """Summarize quality metrics."""
        quality_scores = plan_data.get('quality_scores', {})

        return {
            'overall_score': quality_scores.get('overall_score', 0),
            'completeness': quality_scores.get('completeness_score', 0),
            'consistency': quality_scores.get('consistency_score', 0),
            'feasibility': quality_scores.get('feasibility_score', 0),
            'dependency_validity': quality_scores.get('dependency_score', 0),
            'balance': quality_scores.get('balance_score', 0),
            'meets_thresholds': quality_scores.get('meets_thresholds', False),
            'critical_issues': quality_scores.get('critical_issues', [])
        }

    def _summarize_timeline(self, plan_data: Dict[str, Any]) -> Dict[str, Any]:
        """Summarize timeline information."""
        return {
            'start_date': plan_data.get('created_at', 'Unknown'),
            'estimated_completion': 'Not available',
            'total_duration': 'Not available',
            'milestones_completed': 0,
            'milestones_total': 0
        }

    def _aggregate_data(
        self,
        plans_data: List[Dict[str, Any]],
        group_by: str
    ) -> Dict[str, Any]:
        """Aggregate data from multiple plans."""
        aggregated = {
            'plans': plans_data,
            'summary': {
                'total_plans': len(plans_data),
                'total_sub_problems': sum(len(p.get('sub_problems', [])) for p in plans_data),
                'strategies_used': list(set(p.get('strategy') for p in plans_data)),
                'average_quality': sum(p.get('quality_scores', {}).get('overall_score', 0) for p in plans_data) / len(plans_data) if plans_data else 0
            }
        }

        # Group by specified field
        if group_by == 'date':
            groups = defaultdict(list)
            for plan in plans_data:
                date = plan.get('created_at', 'unknown')[:10]  # Extract date part
                groups[date].append(plan)
            aggregated['groups'] = dict(groups)

        elif group_by == 'strategy':
            groups = defaultdict(list)
            for plan in plans_data:
                strategy = plan.get('strategy', 'unknown')
                groups[strategy].append(plan)
            aggregated['groups'] = dict(groups)

        elif group_by == 'problem_type':
            groups = defaultdict(list)
            for plan in plans_data:
                ptype = plan.get('problem', {}).get('type', 'unknown')
                groups[ptype].append(plan)
            aggregated['groups'] = dict(groups)

        return aggregated

    def _get_default_sections(self) -> List[str]:
        """Get default report sections."""
        return [
            'overview',
            'problem_definition',
            'sub_problems',
            'dependencies',
            'quality_assessment',
            'timeline',
            'recommendations',
            'raw_data' if self.include_raw_data else None
        ]

    def _generate_pdf_report(
        self,
        data: Dict[str, Any],
        template_name: Optional[str],
        output_path: Optional[str],
        sections: List[str]
    ) -> str:
        """Generate PDF report."""
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("PDF generation requires ReportLab. Install with: pip install reportlab")

        logger.info("Generating PDF report")

        output_path = output_path or str(self.output_dir / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

        # Create PDF document
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()

        # Add custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor(self.branding.get('primary_color', '#2C3E50')),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor(self.branding.get('secondary_color', '#34495E')),
            spaceAfter=12
        )

        # Add title
        title = data.get('title', 'Decomposition Report')
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Add metadata
        metadata = data.get('metadata', {})
        story.append(Paragraph(f"<b>Generated:</b> {metadata.get('generated_at', 'Unknown')}", styles['Normal']))
        story.append(Paragraph(f"<b>Plan ID:</b> {metadata.get('plan_id', 'Unknown')}", styles['Normal']))
        story.append(Spacer(1, 0.3 * inch))

        # Add sections
        if 'overview' in sections:
            story.append(Paragraph("Overview", heading_style))
            story.append(self._create_overview_table(data, styles))
            story.append(Spacer(1, 0.2 * inch))

        if 'problem_definition' in sections:
            story.append(Paragraph("Problem Definition", heading_style))
            story.append(self._create_problem_section(data, styles))
            story.append(Spacer(1, 0.2 * inch))

        if 'sub_problems' in sections:
            story.append(Paragraph("Sub-Problems", heading_style))
            story.append(self._create_subproblems_table(data, styles))
            story.append(Spacer(1, 0.2 * inch))

        if 'quality_assessment' in sections:
            story.append(Paragraph("Quality Assessment", heading_style))
            story.append(self._create_quality_section(data, styles))
            story.append(Spacer(1, 0.2 * inch))

        # Build PDF
        doc.build(story)

        logger.info(f"PDF report generated: {output_path}")
        return output_path

    def _create_overview_table(self, data: Dict[str, Any], styles: Dict) -> Table:
        """Create overview table for PDF."""
        metrics = data.get('key_metrics', {})
        metadata = data.get('metadata', {})

        table_data = [
            ['Metric', 'Value'],
            ['Total Sub-Problems', str(metrics.get('total_sub_problems', 0))],
            ['Completed', str(metrics.get('completed', 0))],
            ['In Progress', str(metrics.get('in_progress', 0))],
            ['Pending', str(metrics.get('pending', 0))],
            ['Completion', f"{metrics.get('completion_percentage', 0):.1f}%"],
            ['Average Quality', f"{metrics.get('average_quality_score', 0):.1f}"]
        ]

        table = Table(table_data, colWidths=[2.5 * inch, 2.5 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        return table

    def _create_problem_section(self, data: Dict[str, Any], styles: Dict) -> Paragraph:
        """Create problem section for PDF."""
        problem = data.get('problem', {})
        description = problem.get('description', 'No description available')

        return Paragraph(description, styles['Normal'])

    def _create_subproblems_table(self, data: Dict[str, Any], styles: Dict) -> Table:
        """Create sub-problems table for PDF."""
        sub_problems = data.get('sub_problems', [])

        table_data = [['ID', 'Title', 'Status', 'Quality']]

        for sp in sub_problems[:10]:  # Limit to first 10
            table_data.append([
                sp.get('id', '')[:10],
                sp.get('title', '')[:30],
                sp.get('status', 'unknown'),
                f"{sp.get('quality_score', 0):.1f}"
            ])

        table = Table(table_data, colWidths=[0.8 * inch, 3 * inch, 1 * inch, 0.8 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        return table

    def _create_quality_section(self, data: Dict[str, Any], styles: Dict) -> Table:
        """Create quality assessment section for PDF."""
        quality = data.get('quality_summary', {})

        table_data = [
            ['Dimension', 'Score'],
            ['Overall', f"{quality.get('overall_score', 0):.1f}"],
            ['Completeness', f"{quality.get('completeness', 0):.1f}"],
            ['Consistency', f"{quality.get('consistency', 0):.1f}"],
            ['Feasibility', f"{quality.get('feasibility', 0):.1f}"],
            ['Dependency Validity', f"{quality.get('dependency_validity', 0):.1f}"],
            ['Balance', f"{quality.get('balance', 0):.1f}"]
        ]

        table = Table(table_data, colWidths=[2 * inch, 1.5 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        return table

    def _generate_html_report(
        self,
        data: Dict[str, Any],
        template_name: Optional[str],
        output_path: Optional[str],
        sections: List[str]
    ) -> str:
        """Generate HTML interactive report."""
        logger.info("Generating HTML report")

        output_path = output_path or str(self.output_dir / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")

        # Try to use custom template
        if template_name and self.jinja_env:
            try:
                template = self.jinja_env.get_template(f"{template_name}.html")
                html_content = template.render(data=data, sections=sections)
            except TemplateNotFound:
                logger.warning(f"Template '{template_name}' not found, using default")
                html_content = self._create_default_html(data, sections)
        else:
            html_content = self._create_default_html(data, sections)

        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        logger.info(f"HTML report generated: {output_path}")
        return output_path

    def _create_default_html(self, data: Dict[str, Any], sections: List[str]) -> str:
        """Create default HTML report."""
        metrics = data.get('key_metrics', {})
        metadata = data.get('metadata', {})
        problem = data.get('problem', {})
        sub_problems = data.get('sub_problems', [])
        quality = data.get('quality_summary', {})

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{data.get('title', 'Decomposition Report')}</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            background-color: white;
            border-radius: 8px;
            padding: 30px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: {self.branding.get('primary_color', '#2C3E50')};
            border-bottom: 3px solid {self.branding.get('secondary_color', '#34495E')};
            padding-bottom: 10px;
        }}
        h2 {{
            color: {self.branding.get('secondary_color', '#34495E')};
            margin-top: 30px;
        }}
        .metadata {{
            background-color: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            margin: 20px 0;
        }}
        .metrics {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }}
        .metric-card {{
            background-color: #f8f9fa;
            padding: 20px;
            border-radius: 5px;
            text-align: center;
            border-left: 4px solid {self.branding.get('primary_color', '#2C3E50')};
        }}
        .metric-value {{
            font-size: 2em;
            font-weight: bold;
            color: {self.branding.get('primary_color', '#2C3E50')};
        }}
        .metric-label {{
            color: #666;
            font-size: 0.9em;
            margin-top: 5px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background-color: {self.branding.get('secondary_color', '#34495E')};
            color: white;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .status-completed {{ color: #28a745; font-weight: bold; }}
        .status-in-progress {{ color: #ffc107; font-weight: bold; }}
        .status-pending {{ color: #6c757d; }}
        .status-blocked {{ color: #dc3545; font-weight: bold; }}
        .quality-bar {{
            width: 100%;
            height: 20px;
            background-color: #e0e0e0;
            border-radius: 10px;
            overflow: hidden;
        }}
        .quality-fill {{
            height: 100%;
            background-color: {self.branding.get('primary_color', '#2C3E50')};
            transition: width 0.3s ease;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{data.get('title', 'Decomposition Report')}</h1>

        <div class="metadata">
            <p><strong>Generated:</strong> {metadata.get('generated_at', 'Unknown')}</p>
            <p><strong>Plan ID:</strong> {metadata.get('plan_id', 'Unknown')}</p>
            <p><strong>Strategy:</strong> {metadata.get('strategy', 'Unknown')}</p>
        </div>

        <div class="metrics">
            <div class="metric-card">
                <div class="metric-value">{metrics.get('total_sub_problems', 0)}</div>
                <div class="metric-label">Total Sub-Problems</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{metrics.get('completion_percentage', 0):.1f}%</div>
                <div class="metric-label">Completion</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{metrics.get('average_quality_score', 0):.1f}</div>
                <div class="metric-label">Avg Quality Score</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{metrics.get('completed', 0)}</div>
                <div class="metric-label">Completed</div>
            </div>
        </div>

        <h2>Problem Definition</h2>
        <p><strong>Title:</strong> {problem.get('title', 'Unknown')}</p>
        <p><strong>Description:</strong> {problem.get('description', 'No description available')}</p>

        <h2>Sub-Problems</h2>
        <table>
            <thead>
                <tr>
                    <th>ID</th>
                    <th>Title</th>
                    <th>Status</th>
                    <th>Quality Score</th>
                </tr>
            </thead>
            <tbody>
"""

        for sp in sub_problems:
            status_class = f"status-{sp.get('status', 'pending').replace('_', '-')}"
            html += f"""
                <tr>
                    <td>{sp.get('id', '')}</td>
                    <td>{sp.get('title', '')}</td>
                    <td class="{status_class}">{sp.get('status', 'unknown').replace('_', ' ').title()}</td>
                    <td>{sp.get('quality_score', 0):.1f}</td>
                </tr>
"""

        html += """
            </tbody>
        </table>

        <h2>Quality Assessment</h2>
        <table>
            <thead>
                <tr>
                    <th>Dimension</th>
                    <th>Score</th>
                    <th>Visual</th>
                </tr>
            </thead>
            <tbody>
"""

        for dimension, score in [
            ('Overall', quality.get('overall_score', 0)),
            ('Completeness', quality.get('completeness', 0)),
            ('Consistency', quality.get('consistency', 0)),
            ('Feasibility', quality.get('feasibility', 0)),
            ('Dependency Validity', quality.get('dependency_validity', 0)),
            ('Balance', quality.get('balance', 0))
        ]:
            html += f"""
                <tr>
                    <td>{dimension}</td>
                    <td>{score:.1f}</td>
                    <td>
                        <div class="quality-bar">
                            <div class="quality-fill" style="width: {score}%"></div>
                        </div>
                    </td>
                </tr>
"""

        html += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""
        return html

    def _generate_markdown_report(
        self,
        data: Dict[str, Any],
        template_name: Optional[str],
        output_path: Optional[str],
        sections: List[str]
    ) -> str:
        """Generate Markdown report."""
        logger.info("Generating Markdown report")

        output_path = output_path or str(self.output_dir / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md")

        metrics = data.get('key_metrics', {})
        metadata = data.get('metadata', {})
        problem = data.get('problem', {})
        sub_problems = data.get('sub_problems', [])
        quality = data.get('quality_summary', {})

        md = f"""# {data.get('title', 'Decomposition Report')}

**Generated:** {metadata.get('generated_at', 'Unknown')}
**Plan ID:** {metadata.get('plan_id', 'Unknown')}
**Strategy:** {metadata.get('strategy', 'Unknown')}

## Overview

| Metric | Value |
|--------|-------|
| Total Sub-Problems | {metrics.get('total_sub_problems', 0)} |
| Completed | {metrics.get('completed', 0)} |
| In Progress | {metrics.get('in_progress', 0)} |
| Pending | {metrics.get('pending', 0)} |
| Completion | {metrics.get('completion_percentage', 0):.1f}% |
| Average Quality Score | {metrics.get('average_quality_score', 0):.1f} |

## Problem Definition

**Title:** {problem.get('title', 'Unknown')}

**Description:**
{problem.get('description', 'No description available')}

## Sub-Problems

| ID | Title | Status | Quality Score |
|----|-------|--------|---------------|
"""

        for sp in sub_problems:
            md += f"| {sp.get('id', '')} | {sp.get('title', '')} | {sp.get('status', 'unknown').replace('_', ' ').title()} | {sp.get('quality_score', 0):.1f} |\n"

        md += """
## Quality Assessment

| Dimension | Score |
|-----------|-------|
| Overall | {:.1f} |
| Completeness | {:.1f} |
| Consistency | {:.1f} |
| Feasibility | {:.1f} |
| Dependency Validity | {:.1f} |
| Balance | {:.1f} |
""".format(
            quality.get('overall_score', 0),
            quality.get('completeness', 0),
            quality.get('consistency', 0),
            quality.get('feasibility', 0),
            quality.get('dependency_validity', 0),
            quality.get('balance', 0)
        )

        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(md)

        logger.info(f"Markdown report generated: {output_path}")
        return output_path

    def _generate_excel_report(
        self,
        data: Dict[str, Any],
        output_path: Optional[str],
        sections: List[str]
    ) -> str:
        """Generate Excel report with multiple sheets."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("Excel generation requires openpyxl. Install with: pip install openpyxl")

        logger.info("Generating Excel report")

        output_path = output_path or str(self.output_dir / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        wb = Workbook()

        # Overview sheet
        ws_overview = wb.active
        ws_overview.title = "Overview"
        self._create_excel_overview_sheet(ws_overview, data)

        # Sub-problems sheet
        ws_subproblems = wb.create_sheet("Sub-Problems")
        self._create_excel_subproblems_sheet(ws_subproblems, data)

        # Quality sheet
        ws_quality = wb.create_sheet("Quality")
        self._create_excel_quality_sheet(ws_quality, data)

        # Raw data sheet
        if self.include_raw_data:
            ws_raw = wb.create_sheet("Raw Data")
            self._create_excel_raw_data_sheet(ws_raw, data)

        # Save workbook
        wb.save(output_path)

        logger.info(f"Excel report generated: {output_path}")
        return output_path

    def _create_excel_overview_sheet(self, ws, data: Dict[str, Any]):
        """Create overview sheet in Excel."""
        metrics = data.get('key_metrics', {})
        metadata = data.get('metadata', {})

        # Title
        ws['A1'] = 'Decomposition Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')

        # Metadata
        row = 3
        ws[f'A{row}'] = 'Generated:'
        ws[f'B{row}'] = metadata.get('generated_at', 'Unknown')
        row += 1
        ws[f'A{row}'] = 'Plan ID:'
        ws[f'B{row}'] = metadata.get('plan_id', 'Unknown')
        row += 2

        # Metrics
        ws[f'A{row}'] = 'Metric'
        ws[f'B{row}'] = 'Value'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        for metric, value in [
            ('Total Sub-Problems', metrics.get('total_sub_problems', 0)),
            ('Completed', metrics.get('completed', 0)),
            ('In Progress', metrics.get('in_progress', 0)),
            ('Pending', metrics.get('pending', 0)),
            ('Completion %', metrics.get('completion_percentage', 0)),
            ('Average Quality', metrics.get('average_quality_score', 0))
        ]:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1

        # Formatting
        for col in ['A', 'B']:
            ws.column_dimensions[col].width = 20

    def _create_excel_subproblems_sheet(self, ws, data: Dict[str, Any]):
        """Create sub-problems sheet in Excel."""
        sub_problems = data.get('sub_problems', [])

        # Header
        headers = ['ID', 'Title', 'Status', 'Quality Score', 'Dependencies']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Data
        for row, sp in enumerate(sub_problems, 2):
            ws.cell(row=row, column=1, value=sp.get('id', ''))
            ws.cell(row=row, column=2, value=sp.get('title', ''))
            ws.cell(row=row, column=3, value=sp.get('status', ''))
            ws.cell(row=row, column=4, value=sp.get('quality_score', 0))
            ws.cell(row=row, column=5, value=', '.join(sp.get('dependencies', [])))

        # Formatting
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20

    def _create_excel_quality_sheet(self, ws, data: Dict[str, Any]):
        """Create quality assessment sheet in Excel."""
        quality = data.get('quality_summary', {})

        # Header
        ws['A1'] = 'Quality Assessment'
        ws['A1'].font = Font(size=14, bold=True)

        # Data
        row = 3
        ws[f'A{row}'] = 'Dimension'
        ws[f'B{row}'] = 'Score'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        for dimension, score in [
            ('Overall', quality.get('overall_score', 0)),
            ('Completeness', quality.get('completeness', 0)),
            ('Consistency', quality.get('consistency', 0)),
            ('Feasibility', quality.get('feasibility', 0)),
            ('Dependency Validity', quality.get('dependency_validity', 0)),
            ('Balance', quality.get('balance', 0))
        ]:
            ws[f'A{row}'] = dimension
            ws[f'B{row}'] = score
            row += 1

        # Formatting
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

    def _create_excel_raw_data_sheet(self, ws, data: Dict[str, Any]):
        """Create raw data sheet in Excel."""
        # Simply dump JSON data
        import json
        ws['A1'] = 'Raw Data (JSON)'
        ws['A1'].font = Font(bold=True)

        json_str = json.dumps(data, indent=2, default=str)
        lines = json_str.split('\n')

        for row, line in enumerate(lines, 2):
            ws[f'A{row}'] = line

        ws.column_dimensions['A'].width = 100

    def _generate_json_report(self, data: Dict[str, Any], output_path: Optional[str]) -> str:
        """Generate JSON data export."""
        logger.info("Generating JSON report")

        output_path = output_path or str(self.output_dir / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")

        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)

        logger.info(f"JSON report generated: {output_path}")
        return output_path


# Convenience functions for common reporting tasks
def generate_quick_report(
    decomposition_plan: Any,
    format: str = 'html',
    output_dir: str = './reports'
) -> str:
    """
    Generate a quick report with default settings.

    Args:
        decomposition_plan: DecompositionPlan object
        format: Output format (html, pdf, md, xlsx, json)
        output_dir: Output directory

    Returns:
        str: Path to generated report
    """
    generator = ReportGenerator({'output_dir': output_dir})
    return generator.generate_report(decomposition_plan, output_format=format)


def generate_executive_dashboard(
    decomposition_plans: List[Any],
    output_path: str = './reports/dashboard.html'
) -> str:
    """
    Generate an executive dashboard showing multiple plans.

    Args:
        decomposition_plans: List of DecompositionPlan objects
        output_path: Path for output HTML file

    Returns:
        str: Path to generated dashboard
    """
    generator = ReportGenerator()
    return generator.generate_aggregated_report(
        decomposition_plans,
        output_format='html',
        output_path=output_path
    )
